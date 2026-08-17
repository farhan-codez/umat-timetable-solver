import sys
import shutil
import re
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
SEC_RE = re.compile(r"([A-Za-z]+)(\d+)-([A-Za-z]+)")


def parse_sections(value):
    out = []
    for x in str(value or "").split(","):
        x = x.strip().upper()
        if not x:
            continue
        m = SEC_RE.match(x)
        if not m:
            raise ValueError(f"cannot parse section id {x!r}")
        out.append((m.group(1), m.group(2), m.group(3)))
    return out


def section_ids(r):
    return [f"{p}{lvl}-{sec}" for (p, lvl, sec) in parse_sections(r.get("sections"))]


def is_combined(r):
    secs = parse_sections(r.get("sections"))
    return len({p for (p, _, _) in secs}) > 1


def _num(v):
    try:
        return float(v) if v not in (None, "") else 0.0
    except (TypeError, ValueError):
        return 0.0


def split_shared(sem):
    data_dir = ROOT / "data" / "semesters" / sem
    src = data_dir / "courses.xlsx"
    bak = data_dir / "courses.presplit.xlsx"
    if not bak.exists():
        shutil.copyfile(src, bak)
        print(f"{sem}: backed up -> {bak.name}")

    wb = load_workbook(src)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    rows = [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2) if r[0].value]

    combined = [r for r in rows if is_combined(r)]
    existing = [r for r in rows if not is_combined(r)]
    out = [dict(r) for r in existing]

    merges = []
    news = []

    for r in combined:
        code = str(r["course_code"]).strip()
        num = code.split()[-1]
        by_prog = {}
        for (prog, lvl, sec) in parse_sections(r.get("sections")):
            by_prog.setdefault(prog, []).append(f"{prog}{lvl}-{sec}")

        for prog, sec_list in by_prog.items():
            new_code = f"{prog} {num}"
            want = set(sec_list)
            part = dict(r)
            part["course_code"] = new_code
            part["programme"] = prog
            part["sections"] = ",".join(sorted(want))
            part["cohort"] = None
            part["size"] = None
            part["split"] = "no"
            part["sessions_per_week"] = None

            target = None
            for x in existing:
                if (str(x.get("course_code")) == new_code
                        and str(x.get("programme")) == prog
                        and set(section_ids(x)) == want):
                    target = x
                    break

            if target is not None:
                old_lec, old_prac = _num(target["lecture_hours"]), _num(target["practical_hours"])
                new_lec = old_lec + _num(part["lecture_hours"])
                new_prac = old_prac + _num(part["practical_hours"])
                merges.append((new_code, prog, sorted(want), old_lec, new_lec, old_prac, new_prac,
                               str(target.get("online")), str(part.get("online")),
                               str(target.get("field_work")), str(part.get("field_work"))))
                target["lecture_hours"] = new_lec
                target["practical_hours"] = new_prac
                target["credits"] = max(_num(target["credits"]), _num(part["credits"]))
                target["sessions_per_week"] = None
            else:
                news.append((new_code, prog, sorted(want),
                             _num(part["lecture_hours"]), _num(part["practical_hours"]),
                             str(part.get("online")), str(part.get("field_work"))))
                out.append(part)

    out.sort(key=lambda r: (str(r.get("programme")), str(r.get("level")), str(r.get("course_code"))))

    wout = Workbook()
    wo = wout.active
    wo.append(hdr)
    for r in out:
        wo.append([r.get(c) for c in hdr])
    wout.save(src)

    print(f"{sem}: {len(rows)} rows -> {len(out)} | combined removed: {len(combined)}")
    print(f"  new split rows: {len(news)}")
    for code, prog, secs, lec, prac, onl, fw in sorted(news):
        print(f"    NEW {code:10s} {prog:3s} {','.join(secs):24s} lec={lec:g} prac={prac:g} online={onl} fw={fw}")
    print(f"  merged into existing: {len(merges)}")
    for code, prog, secs, olec, nlec, oprac, nprac, tonl, ponl, tfw, pfw in sorted(merges):
        flag = ""
        if (tonl != ponl) or (tfw != pfw):
            flag = f"  <-- ONLINE/FW MISMATCH existing({tonl}/{tfw}) part({ponl}/{pfw})"
        print(f"    MERGE {code:10s} {prog:3s} {','.join(secs):24s} lec {olec:g}->{nlec:g} prac {oprac:g}->{nprac:g}{flag}")


if __name__ == "__main__":
    for sem in sys.argv[1:] or ("sem1", "sem2"):
        split_shared(sem)

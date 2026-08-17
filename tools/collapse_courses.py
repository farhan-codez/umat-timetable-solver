import sys
import shutil
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from openpyxl import Workbook, load_workbook
from collections import defaultdict
from src.loaders import _canonical_lecturer

ROOT = Path(__file__).resolve().parent.parent
COLS = ['course_code', 'course_name', 'programme', 'level', 'cohort', 'lecturer',
        'lecture_hours', 'practical_hours', 'credits', 'online', 'field_work', 'hours_per_session',
        'sessions_per_week', 'min_room_size', 'sections', 'size']


def sections_of(r):
    s = str(r['sections'] or '').strip()
    if s:
        return {x.strip().upper() for x in s.split(',') if x.strip()}
    target = f"{r['programme']}{r['level']}-"
    c = str(r['cohort']).upper()
    if c == 'AB':
        return {target + 'A', target + 'B'}
    return {target + c}


def collapse(sem):
    data_dir = ROOT / "data" / "semesters" / sem
    src = data_dir / "courses.xlsx"
    bak = data_dir / "courses.backup.xlsx"
    if not bak.exists():
        shutil.copyfile(src, bak)
        print(f"{sem}: backed up -> {bak.name}")

    wb = load_workbook(data_dir / "cohorts.xlsx")
    wc = wb["Sheet1"]
    hc = [c.value for c in wc[1]]
    full_pl = defaultdict(set)
    for r in wc.iter_rows(min_row=2):
        if not r[0].value:
            continue
        d = dict(zip(hc, [c.value for c in r]))
        full_pl[(d['programme'], str(d['level']))].add(f"{d['programme']}{d['level']}-{d['section']}")

    wb2 = load_workbook(src)
    ws = wb2["Sheet1"]
    hdr = [c.value for c in ws[1]]
    rows = [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2) if r[0].value]

    comp_attrs = ('programme', 'level', 'lecture_hours', 'practical_hours',
                  'credits', 'online', 'field_work', 'hours_per_session', 'sessions_per_week', 'min_room_size')
    groups = defaultdict(list)
    for r in rows:
        key = (str(r['course_code']), tuple(str(r[k]) for k in comp_attrs), _canonical_lecturer(r['lecturer']))
        groups[key].append(r)

    out = []
    kept_explicit = set()
    for (code, comp, lec), grp in groups.items():
        prog = str(grp[0]['programme'])
        lev = str(grp[0]['level'])
        full = full_pl.get((prog, lev), set())
        union = set()
        for r in grp:
            union |= sections_of(r)
        same_prog = all(s.startswith(prog + lev + '-') for s in union)
        name = next((str(r['course_name']) for r in grp if str(r['course_name']).strip()), '')
        if same_prog and union == full:
            base = dict(grp[0])
            base['lecturer'] = lec
            base['course_name'] = name
            base['cohort'] = 'AB' if len(full) > 1 else next(iter(full)).split('-')[-1]
            base['sections'] = ''
            out.append(base)
        else:
            kept_explicit.add(code)
            for r in grp:
                r['lecturer'] = lec
                if not str(r['course_name']).strip():
                    r['course_name'] = name
                out.append(dict(r))

    out.sort(key=lambda r: (str(r['programme']), str(r['level']), str(r['course_code'])))

    wout = Workbook()
    wo = wout.active
    wo.append(COLS)
    for r in out:
        wo.append([r.get(c) for c in COLS])
    wout.save(src)
    print(f"{sem}: {len(rows)} rows -> {len(out)} | kept-explicit codes: {len(kept_explicit)}")
    print('  kept explicit:', sorted(kept_explicit))


if __name__ == "__main__":
    for sem in ("sem1", "sem2"):
        collapse(sem)

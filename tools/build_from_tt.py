"""Build data/semesters/<sem>/courses.xlsx from the extracted timetable.

Each course becomes ONE simple row (course code, name, programme, level,
lecturer, weekly hours, online/field-work flags, session pattern). Sections
are NOT entered by the admin: the loader derives the sections from the
programme+level and auto-splits the class into per-section streams when the
combined class is too large (size > 90). Courses the original timetable
delivered as a single combined AB class are marked split=no so they keep their
hours as one class.

This replaces the old tools/collapse_courses.py, which merged A/B rows but kept
only one section's hours (halving the teaching load).
"""
import shutil
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from openpyxl import Workbook, load_workbook
from src.loaders import _canonical_lecturer

ROOT = Path(__file__).resolve().parent.parent
COLS = ['course_code', 'course_name', 'programme', 'level', 'cohort', 'lecturer',
        'lecture_hours', 'practical_hours', 'credits', 'online', 'field_work',
        'hours_per_session', 'sessions_per_week', 'split', 'min_room_size', 'sections', 'size']


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _int(v, default=1):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _truthy(v):
    return str(v).strip().lower() in ("yes", "y", "1", "true", "online")


def secs_of(d):
    s = str(d.get('sections') or '').strip()
    if s:
        return frozenset(x.strip().upper() for x in s.split(',') if x.strip())
    return frozenset()


def load_tt(path):
    wb = load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, row))
        if d.get('course_code'):
            out.append(d)
    return out


def load_cohorts(sem):
    wb = load_workbook(ROOT / f"data/semesters/{sem}/cohorts.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    cohorts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, row))
        if d.get('programme'):
            cohorts[f"{str(d['programme']).strip()}{int(d['level'])}-{str(d['section']).strip().upper()}"] = int(d['size'])
    return cohorts


def build(sem, tt_path):
    rows = load_tt(tt_path)
    cohorts = load_cohorts(sem)

    # 1) partition per course: physical, online, fieldwork
    per_code = defaultdict(list)
    for d in rows:
        per_code[str(d.get('course_code')).strip()].append(d)

    out = []
    for code in sorted(per_code):
        grp = per_code[code]
        name = next((str(d.get('course_name')) for d in grp if str(d.get('course_name') or '').strip()), '')
        prog = str(grp[0].get('programme') or '').strip()
        lev = int(_num(grp[0].get('level')))
        credits = max(_num(d.get('credits')) for d in grp)
        all_of_pl = {sid for sid in cohorts if sid.startswith(f"{prog}{lev}-")}

        for d in grp:
            dur = _int(d.get('hours_per_session'), 2)
            spw = _int(d.get('sessions_per_week'), 1)
            online = _truthy(d.get('online'))
            fw = _truthy(d.get('field_work'))
            lect_h = _num(d.get('lecture_hours'))
            prac_h = _num(d.get('practical_hours'))
            lec = _canonical_lecturer(d.get('lecturer'))
            sset = secs_of(d)
            size = str(d.get('size') or '').strip()

            base = {
                'course_code': code, 'course_name': name, 'programme': prog, 'level': lev,
                'cohort': '', 'lecturer': lec, 'lecture_hours': lect_h, 'practical_hours': prac_h,
                'credits': credits, 'online': 'yes' if online else 'no',
                'field_work': 'yes' if fw else 'no', 'hours_per_session': dur,
                'sessions_per_week': spw, 'split': '', 'min_room_size': 0, 'sections': '',
                'size': size,
            }

            if sset:
                base['sections'] = ','.join(sorted(sset))
                # multi-section or online / fieldwork deliveries are never split;
                # a single-section row cannot be split anyway
                if online or fw or len(sset) > 1:
                    base['split'] = 'no'
            else:
                # no sections on the extracted row (fallback): keep the whole
                # programme+level together as one combined class
                base['sections'] = ','.join(sorted(all_of_pl))
                base['split'] = 'no'
            out.append(base)

    out.sort(key=lambda r: (r['programme'], r['level'], r['course_code']))
    return out


def main():
    for sem, sub in [('sem1', '_sem1_tt'), ('sem2', '_sem2_tt')]:
        src = ROOT / f"data/semesters/{sem}/courses.xlsx"
        bak = ROOT / f"data/semesters/{sem}/courses.collapsed.xlsx"
        if not bak.exists():
            shutil.copyfile(src, bak)
            print(f"{sem}: backed up current (collapsed) file -> {bak.name}")

        rows = build(sem, ROOT / f"output/{sub}/courses.xlsx")
        wout = Workbook()
        wo = wout.active
        wo.title = 'Sheet1'
        wo.append(COLS)
        for r in rows:
            wo.append([r.get(c) for c in COLS])
        wout.save(src)
        print(f"{sem}: wrote {len(rows)} course rows -> courses.xlsx")


if __name__ == "__main__":
    main()

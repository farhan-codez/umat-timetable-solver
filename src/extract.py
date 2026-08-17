"""Extract courses and rooms from the raw SRID teaching timetable into the solver's data files.

Usage:
    .venv\\Scripts\\python.exe -m src.extract

Reads the first .xlsx in data/input, parses every class cell, and writes:
    data/courses.xlsx   - one row per course offering (code, programme, level, cohort, hours)
    data/rooms.xlsx     - rooms found in the timetable (with capacity hints)
    output/extraction_review.xlsx - full parse results + source cells for manual checking

This never touches data/cohorts.xlsx.
"""

import argparse
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
INPUT_DIR = ROOT / "data" / "input"
DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "output"

TEACH_COLS = list(range(2, 8)) + list(range(9, 15))  # B..G morning, I..N afternoon

FIELD_ROOMS = {"FIELD WORK", "FIELD WORK/ LAB WORK", "FIELD WORK / LAB WORK"}
LAB_ROOMS = {"COMPUTER LAB"}
ONLINE_ROOMS = {"ONLINE", "SR 2", "SR 6", "HARDWARE LAB"}  # SR 2 / SR 6 / HARDWARE LAB are online placeholders, not real rooms

DEFAULT_ROOM_CAP = {"M. AUDITORIUM": 120,
                    "COMPUTER LAB": 80, "HARDWARE LAB": 80}

COURSE_HEADERS = [
    "course_code", "course_name", "programme", "level", "cohort",
    "lecturer", "lecture_hours", "practical_hours", "credits",
    "online", "field_work", "hours_per_session", "sessions_per_week", "min_room_size",
    "sections", "size",
]
ROOM_HEADERS = ["name", "capacity", "kind"]

HEADER_FILL = PatternFill("solid", fgColor="0D9488")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def find_timetable():
    candidates = [p for p in INPUT_DIR.glob("*.xlsx") if not p.name.startswith("~$")]
    if not candidates:
        sys.exit(f"No timetable found in {INPUT_DIR}")
    if len(candidates) > 1:
        print(f"Multiple files in {INPUT_DIR}, using {candidates[0].name}")
    return candidates[0]


def merged_map(ws):
    mm = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    mm[(r, c)] = True
    return mm


def cell_span(ws, r, c):
    for rng in ws.merged_cells.ranges:
        if rng.min_row == r and rng.min_col == c and rng.max_row == r:
            return rng.max_col - rng.min_col + 1
    return 1


def _start_time(col):
    if 2 <= col <= 7:
        return f"{6 + (col - 2):02d}:30"
    if 9 <= col <= 14:
        return f"{13 + (col - 9):02d}:00"
    return "?"


def collect_records(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    records = []
    for day in wb.sheetnames:
        ws = wb[day]
        mm = merged_map(ws)
        cur = None
        for r in range(9, ws.max_row + 1):
            v = ws.cell(row=r, column=1).value
            if v and str(v).strip():
                cur = re.sub(r"\s+", " ", str(v).strip())
            if not cur:
                continue
            for c in TEACH_COLS:
                if (r, c) in mm:
                    continue
                cell = ws.cell(row=r, column=c)
                if cell.value is None or not str(cell.value).strip():
                    continue
                records.append({
                    "day": day,
                    "room": cur,
                    "span": cell_span(ws, r, c),
                    "time": _start_time(c),
                    "text": re.sub(r"\s+", " ", str(cell.value).strip()),
                })
    return records


def venue_kind(room):
    r = room.strip().upper()
    if r in LAB_ROOMS:
        return "lab"
    if r in FIELD_ROOMS:
        return "field"
    if r in ONLINE_ROOMS:
        return "online"
    return "lecture"


def detect_section(tail):
    t = tail.strip()
    while t.startswith("("):
        end = t.find(")")
        if end == -1:
            break
        t = t[end + 1:].lstrip()
    m = re.match(r"A\s*&\s*B", t)
    if m:
        return "AB", t[m.end():].strip()
    m = re.match(r"A(?!\.)(?=\s|$)", t)
    if m:
        return "A", t[1:].strip()
    m = re.match(r"B(?!\.)(?=\s|$)", t)
    if m:
        return "B", t[1:].strip()
    return None, t


def clean_lecturer(tail, section):
    t = tail.strip()
    if section:
        t = re.sub(r"^(A\s*&\s*B|A&B)\b", "", t).strip()
        t = re.sub(r"^(A|B)(?!\.)\b", "", t).strip()
    t = re.sub(r"\([^)]*\)", "", t)
    t = re.sub(r"&", "", t)
    t = re.sub(r"\s+", " ", t).strip(" -")
    return t


def parse_part(part_text, pending, raw):
    text = part_text.strip()
    if not text:
        return [], pending, []
    if re.fullmatch(r"[A-Z]{1,3}", text):
        return [], pending + [text], []
    m = re.search(r"([A-Z]{1,3})(?:\s*\([^)]*\))*\s*(\d{2,3})", text)
    if not m:
        return [], pending, [("unparsed-fragment", raw, text)]
    pre, num = m.group(1), int(m.group(2))
    prefixes = pending + [pre]
    tail = text[m.end():]
    section, rest = detect_section(tail)
    lecturer = clean_lecturer(rest, section)
    p_marker = bool(re.search(r"\(P\s*\)", text, re.IGNORECASE))

    offers = []
    for prefix in prefixes:
        offers.append({"code": f"{prefix} {num}", "prefix": prefix, "prog": prefix,
                       "level": (num // 100) * 100, "section": section,
                       "lecturer": lecturer, "p_marker": p_marker})
    return offers, [], []


def parse_cell(rec):
    parts = rec["text"].split("/")
    pending = []
    offers = []
    issues = []
    for part in parts:
        part_offers, pending, part_issues = parse_part(part, pending, rec["text"])
        offers.extend(part_offers)
        issues.extend(part_issues)
    if pending:
        issues.append(("trailing-prefix", rec["text"], str(pending)))
    return offers, issues


def load_sections(data_dir):
    df = pd.read_excel(data_dir / "cohorts.xlsx")
    sections = defaultdict(list)
    for _, row in df.iterrows():
        prog = row.get("programme")
        if prog is None or (isinstance(prog, float) and pd.isna(prog)):
            continue
        key = (str(prog).strip().upper(), int(row["level"]))
        sections[key].append(str(row["section"]).strip().upper())
    return {k: sorted(v) for k, v in sections.items()}


def resolve_cohort(secs, tag):
    if len(secs) == 2:
        return tag if tag in ("A", "B", "AB") else "AB"
    single = secs[0]
    if tag in (None, "AB") or tag == single:
        return single
    return single


def write_xlsx(path, sheet, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, row in enumerate(rows, start=2):
        for col, h in enumerate(headers, start=1):
            ws.cell(row=i, column=col, value=row.get(h))
    wb.save(path)
    print(f"Wrote {path.name} ({len(rows)} rows)")


def write_review(path, course_rows, cell_rows, anomalies):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Courses"
    headers = COURSE_HEADERS + ["weekly_hours", "blocks"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, r in enumerate(course_rows, start=2):
        vals = [r.get(h) for h in COURSE_HEADERS]
        vals += [r["lecture_hours"] + r["practical_hours"], r["sessions_per_week"]]
        for col, v in enumerate(vals, start=1):
            ws.cell(row=i, column=col, value=v)

    ws2 = wb.create_sheet("Cells")
    headers2 = ["day", "time", "room", "hours", "text", "code", "programme", "level",
                "cohort", "lecturer", "practical", "no_room"]
    for col, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for i, r in enumerate(cell_rows, start=2):
        for col, h in enumerate(headers2, start=1):
            ws2.cell(row=i, column=col, value=r.get(h))

    ws3 = wb.create_sheet("Anomalies")
    for col, h in enumerate(["kind", "raw", "detail"], start=1):
        ws3.cell(row=1, column=col, value=h).font = Font(bold=True)
    for i, (kind, raw, extra) in enumerate(anomalies, start=2):
        ws3.cell(row=i, column=1, value=kind)
        ws3.cell(row=i, column=2, value=raw)
        ws3.cell(row=i, column=3, value=extra)

    for ws_name, widths in (("Courses", [14, 24, 12, 8, 8, 24, 12, 14, 8, 8, 16, 12, 12, 12, 6]),
                            ("Cells", [9, 6, 16, 6, 46, 10, 12, 8, 8, 22, 8, 8])):
        for col, w in enumerate(widths, start=1):
            wb[ws_name].column_dimensions[chr(64 + col)].width = w
    wb.save(path)
    print(f"Wrote review {path.name}")


def list_timetables(input_dir):
    """All teaching-timetable workbooks in data/input: name matches
    ('TEACHING TIMETABLE') first, then workbooks with weekday-named sheets."""
    candidates = [p for p in Path(input_dir).glob("*.xlsx") if not p.name.startswith("~$")]
    if not candidates:
        return []
    named = sorted(p for p in candidates if "TEACHING TIMETABLE" in p.name.upper())
    others = []
    for p in sorted(candidates):
        if p.name in {q.name for q in named}:
            continue
        try:
            wb = openpyxl.load_workbook(p, read_only=True)
            names = wb.sheetnames
            wb.close()
        except Exception:
            continue
        if any(re.fullmatch(r"(MON|TUE|WED|THU|FRI|SAT|SUN)(DAY)?", n.strip().upper()) for n in names):
            others.append(p)
    return named + others


def detect_timetable(input_dir):
    """Find the assembled teaching-timetable workbook (day-named sheets or
    'TEACHING TIMETABLE' in the name) among the files in data/input."""
    tts = list_timetables(input_dir)
    return tts[0] if tts else None


def extract_timetable(path, data_dir, output_dir=None):
    """Parse one teaching-timetable workbook into courses.xlsx / rooms.xlsx.
    Returns a dict with counts and anomalies for the caller to report."""
    print(f"Reading {path.name}")
    records = collect_records(path)
    sections = load_sections(data_dir)
    cohort_size = {}
    try:
        cdf = pd.read_excel(data_dir / "cohorts.xlsx")
        for _, r in cdf.iterrows():
            if r.get("programme") is None or (isinstance(r.get("programme"), float) and pd.isna(r.get("programme"))):
                continue
            cohort_size[f"{str(r['programme']).strip().upper()}{int(r['level'])}-{str(r['section']).strip().upper()}"] = int(r["size"])
    except Exception:
        pass
    print(f"Cells parsed from timetable: {len(records)}")

    cell_rows = []
    anomalies = []
    ok = []          # (rec, offer-dict, cohort)
    skipped = []     # (rec, offer-dict, reason)
    room_caps = defaultdict(set)
    stats = Counter()

    for rec in records:
        offers, issues = parse_cell(rec)
        anomalies.extend(issues)
        if not offers:
            stats["unparsed-cells"] += 1
            continue

        vk = venue_kind(rec["room"])
        practical = vk in ("lab", "field") or any(o["p_marker"] for o in offers)
        no_room = vk in ("online", "field")
        is_vle = re.search(r"\(VLE\)", rec["text"], re.IGNORECASE) is not None
        field_work = vk == "field" and not is_vle
        if is_vle:
            no_room = True  # VLE-marked classes are delivered online - they need no room
        room = re.sub(r"\s*\(\d+\)", "", rec["room"]).strip()
        room_caps[room].update(int(x) for x in re.findall(r"\((\d+)\)", rec["room"]))

        resolved = []
        for offer in offers:
            if offer["prefix"] == "RT":
                stats["skipped-rt"] += 1
                anomalies.append(("rt-skipped", rec["text"], "RT courses skipped by choice"))
                continue
            prog, level = offer["prog"], offer["level"]
            secs = sections.get((prog, level))
            if secs is None:
                skipped.append((rec, offer, f"no cohort {prog}{level} in cohorts.xlsx"))
                stats["skipped-missing-cohort"] += 1
                continue
            cohort = resolve_cohort(secs, offer["section"])
            if offer["section"] not in (None, "AB") and offer["section"] != cohort:
                anomalies.append(("section-mismatch",
                                  rec["text"], f"{prog}{level} tag={offer['section']} -> {cohort}"))
                stats["section-mismatch"] += 1
            section_ids = {f"{prog}{level}-{sec}" for sec in secs if cohort == "AB" or sec == cohort}
            resolved.append((offer, cohort, section_ids))

        if not resolved:
            continue

        # One cell is one teaching block.  Cells naming several courses (e.g.
        # "MC 458 / EL 458" or "CE 151 B / TM 151") are co-taught: the whole
        # class is taught ONCE in this one room to all programmes' students
        # together.  Merge such offers into a single row so the room is only
        # counted once (the class still appears in every programme's timetable
        # through its section list).  A single offer keeps its own row.
        if len(resolved) > 1:
            counts = Counter(o["lecturer"] for o, _, _ in resolved)
            primary = resolved[0]
            offer = dict(primary[0])
            offer["lecturer"] = counts.most_common(1)[0][0]
            cohort = primary[1]
            section_ids = set().union(*(sid for _, _, sid in resolved))
            merged_name = "/".join(sorted({o["code"] for o, _, _ in resolved}))
        else:
            offer, cohort, section_ids = resolved[0]
            merged_name = offer["code"]

        ok.append((rec, offer, cohort, section_ids, practical, no_room, field_work, merged_name))
        cell_rows.append({
            "day": rec["day"], "time": rec["time"], "room": rec["room"], "hours": rec["span"],
            "text": rec["text"], "code": offer["code"], "programme": offer["prog"],
            "level": offer["level"], "cohort": cohort, "lecturer": offer["lecturer"],
            "practical": "P" if (practical or offer["p_marker"]) else "",
            "no_room": "yes" if no_room else "",
        })
        stats["ok-cells"] += 1

    # ---- aggregate into course rows (one row per weekly block; mixed block
    #      lengths become separate rows, each with its own hours_per_session) ----
    agg = defaultdict(lambda: {"cells": [], "lect": Counter(), "practical": 0, "no_room": False, "field_work": False, "sections": set(), "name": ""})
    for rec, offer, cohort, section_ids, is_practical, no_room, field_work, merged_name in ok:
        key = (offer["code"], offer["prog"], offer["level"], cohort, rec["span"], field_work, is_practical, no_room)
        a = agg[key]
        a["cells"].append((rec, is_practical, no_room))
        a["lect"][offer["lecturer"]] += 1
        a["no_room"] = no_room or a["no_room"]
        a["field_work"] = field_work or a["field_work"]
        a["sections"].update(section_ids)
        a["name"] = merged_name or a["name"]
        if is_practical:
            a["practical"] += rec["span"]

    course_rows = []
    for key, a in agg.items():
        code, prog, level, cohort, span, field_work, _, no_room = key
        hours = sum(c["span"] for c, _, _ in a["cells"])
        practical = a["practical"]
        lecturer = a["lect"].most_common(1)[0][0] if a["lect"] else ""
        # cells are grouped by venue kind (online / field / physical / practical),
        # so a course taught partly online, partly in the field or in the lab
        # becomes several rows
        # A class is sized by the smallest room the school actually used for it,
        # so the model reproduces the real attendance: nominal enrolment
        # regularly exceeds the room that is actually used (an A+B class of 116
        # in an 80-seat room, MA400 (85) in SR 4 (40)).  A bigger room in one
        # slot was just spare space - never inflate the size because of it, and
        # never exceed the nominal size.  Online / fieldwork rows need no size.
        size = ""
        if not a["no_room"] and not field_work:
            nominal = sum(cohort_size.get(sid, 0) for sid in a["sections"])
            room_caps_used = set()
            for (rec, _, no_room_c) in a["cells"]:
                if no_room_c:
                    continue
                room = re.sub(r"\s*\(\d+\)", "", rec["room"]).strip()
                caps = [c for c in room_caps[room] if c > 0]
                room_caps_used.add(max(caps) if caps else (DEFAULT_ROOM_CAP.get(room, 0) or 80))
            if nominal and room_caps_used:
                size = min(nominal, min(room_caps_used))
        course_rows.append({
            "course_code": code, "course_name": a["name"], "programme": prog, "level": level,
            "cohort": cohort, "lecturer": lecturer, "lecture_hours": hours - practical,
            "practical_hours": practical, "credits": "",
            "online": "yes" if a["no_room"] and not field_work else "no",
            "field_work": "yes" if field_work else "no",
            "hours_per_session": span, "sessions_per_week": len(a["cells"]),
            "min_room_size": "",
            "sections": ",".join(sorted(a["sections"])),
            "size": size,
        })
    course_rows.sort(key=lambda r: (r["programme"], r["level"], r["cohort"], r["course_code"]))

    # ---- write courses.xlsx (valid rows only) ----
    write_xlsx(data_dir / "courses.xlsx", "Courses", COURSE_HEADERS, course_rows)

    # ---- write rooms.xlsx ----
    room_rows = []
    for room in sorted(room_caps):
        if room in ONLINE_ROOMS or room in FIELD_ROOMS:
            continue
        caps = [c for c in room_caps[room] if c > 0]
        cap = max(caps) if caps else DEFAULT_ROOM_CAP.get(room, 0)
        if cap == 0:
            anomalies.append(("room-capacity-unknown", room, "defaulted to 80"))
            cap = 80
        room_rows.append({"name": room, "capacity": cap, "kind": "lab" if room in LAB_ROOMS else "lecture"})
    room_rows.sort(key=lambda r: (r["kind"] != "lecture", r["name"]))
    write_xlsx(data_dir / "rooms.xlsx", "Rooms", ROOM_HEADERS, room_rows)

    # ---- review workbook ----
    review_path = None
    if output_dir is not None:
        review_path = Path(output_dir) / "extraction_review.xlsx"
        write_review(review_path, course_rows, cell_rows, anomalies)

    return {
        "file": str(path.name),
        "cells": len(records),
        "ok_cells": stats["ok-cells"],
        "unparsed_cells": stats.get("unparsed-cells", 0),
        "courses": len(course_rows),
        "rooms": len(room_rows),
        "skipped_missing_cohort": stats.get("skipped-missing-cohort", 0),
        "skipped_rt": stats.get("skipped-rt", 0),
        "anomaly_counts": dict(Counter(k for k, _, _ in anomalies)),
        "anomalies": [f"[{k}] {extra or raw}" for k, raw, extra in anomalies[:40]],
        "review": str(review_path) if review_path else None,
    }


def _print_summary(result):
    print("\n===== EXTRACTION SUMMARY =====")
    print(f"parsed cells             : {result['ok_cells']}")
    print(f"unparsed cells           : {result['unparsed_cells']}")
    print(f"course rows written      : {result['courses']}")
    print(f"cells skipped (no cohort): {result['skipped_missing_cohort']}")
    print(f"skipped RT (by choice)   : {result['skipped_rt']}")
    print(f"rooms written            : {result['rooms']}")
    print("Anomalies by kind:", result["anomaly_counts"])
    for a in result["anomalies"]:
        print(" ", a)


def main():
    parser = argparse.ArgumentParser(description="Extract courses/rooms from the SRID timetable")
    parser.add_argument("--data-dir", type=Path, default=DATA_DIR, help="output dir for courses.xlsx / rooms.xlsx")
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR, help="dir for the review workbook")
    args = parser.parse_args()

    path = find_timetable()
    result = extract_timetable(path, args.data_dir, args.output_dir)
    _print_summary(result)
    return 0


if __name__ == "__main__":
    sys.exit(main())

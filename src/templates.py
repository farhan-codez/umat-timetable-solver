from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="0D9488")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write(path, sheet_name, headers, widths, n_blank, instructions):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in range(2, 2 + n_blank):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col)
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ins = wb.create_sheet("Instructions")
    for i, line in enumerate(instructions, start=1):
        cell = ins.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True)
    ins.column_dimensions["A"].width = 120
    wb.save(path)


def generate_templates(data_dir):
    data_dir = Path(data_dir)
    data_dir.mkdir(parents=True, exist_ok=True)

    _write(
        data_dir / "rooms.xlsx",
        "Rooms",
        ["name", "capacity", "kind"],
        [16, 12, 10],
        13,
        [
            "ROOMS - one row per room.",
            "",
            "name     : room code, e.g. SR1, SR5A, LH1, Auditorium",
            "capacity : maximum number of students, e.g. 40, 80, 120",
            "kind     : 'lecture' or 'lab' (practical-heavy courses prefer lab rooms,",
            "           then fall back to any room big enough if no lab exists)",
            "",
            "Example row:  SR1 | 40 | lecture",
        ],
    )

    _write(
        data_dir / "cohorts.xlsx",
        "Cohorts",
        ["programme", "level", "section", "size"],
        [14, 8, 10, 8],
        20,
        [
            "COHORTS - one row per A/B section of every programme-year group.",
            "",
            "programme : programme code, e.g. CE, EE, ME",
            "level     : year group, e.g. 100, 200",
            "section   : 'A' or 'B'",
            "size      : number of students in that section",
            "",
            "Rules:",
            "- A and B combined must be <= 120 so an AB (combined) session fits a 120 room.",
            "- If a programme-year is not split, use one section (A) with its real size.",
            "",
            "Example row:  CE | 100 | A | 58",
        ],
    )

    _write(
        data_dir / "courses.xlsx",
        "Courses",
        [
            "course_code", "course_name", "programme", "level", "cohort",
            "lecturer", "lecture_hours", "practical_hours", "credits",
            "online", "field_work", "hours_per_session", "sessions_per_week", "min_room_size",
        ],
        [14, 34, 14, 8, 8, 22, 13, 14, 9, 8, 10, 16, 17, 13],
        50,
        [
            "COURSES - one row per course-offering. A course taught to A and B",
            "separately needs two rows (cohort A and cohort B).",
            "",
            "course_code       : e.g. CE 101",
            "course_name       : e.g. Introduction to Programming",
            "programme         : one programme per row, e.g. CE. For a course shared",
            "                    by several programmes, repeat the row per programme.",
            "level             : year group, e.g. 100",
            "cohort            : 'A', 'B' or 'AB' (AB = combined A+B in one session)",
            "lecturer          : lecturer name (the solver keeps them conflict-free)",
            "lecture_hours     : teaching hours per week (TPC)",
            "practical_hours   : practical hours per week (TPC)",
            "credits           : course credits (TPC)",
            "online            : 'yes' = no room needed (usually an AB combined course),",
            "                    'no' = needs a physical room",
            "field_work        : 'yes' = off-campus field work (Fieldtrip, Engineering",
            "                    Practice, Survey Camp). Also needs no room. A course",
            "                    cannot be both online and field work.",
            "hours_per_session : hours each session lasts (1-12), e.g. 2 for a 2-hour",
            "                    class. Blank = 2.",
            "sessions_per_week : OPTIONAL override; leave blank to auto-compute as",
            "                    (lecture_hours + practical_hours) / hours_per_session",
            "min_room_size     : OPTIONAL smallest room that may be used; blank = auto",
            "",
            "Example row (A-only, on campus):  CE 101 | Intro to Programming | CE | 100 | A | Dr. K. Mensah | 2 | 2 | 3 | no | 2 | |",
            "Example row (combined, online):   MA 121 | Engineering Maths I | CE | 100 | AB | Prof. A. Owusu | 2 | 0 | 2 | yes | 2 | |",
        ],
    )

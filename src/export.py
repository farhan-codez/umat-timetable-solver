from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import json

from .slots import DAYS, N_SLOTS, SLOTS_PER_DAY, SLOT_TIMES, day_index_of, slot_in_day
from .solver import FIELD_WORK_ROOM, ONLINE_ROOM, _is_no_room

HEADER_FILL = PatternFill("solid", fgColor="0D9488")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ONLINE_FILL = PatternFill("solid", fgColor="CFF3EF")
FIELD_FILL = PatternFill("solid", fgColor="D1FADF")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _sheet_names(book, base, existing):
    base = base[:31]
    name = base
    suffix = 2
    while name in existing:
        name = f"{base[:28]} {suffix}"
        suffix += 1
    existing.add(name)
    return book.create_sheet(title=name)


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "B2"


def _room_text(a):
    if a.room == ONLINE_ROOM:
        return "ONLINE"
    if a.room == FIELD_WORK_ROOM:
        return "FIELD WORK"
    return a.room


def _section_label(a):
    """Human-readable section label for one class: 'A', 'B', 'AB' when the
    class serves all sections of one programme-level group, comma-joined
    section ids (e.g. 'CE100-B,TM100-A') for attached cross-programme classes.
    Falls back to the course cohort column for legacy data."""
    secs = sorted(a.session.sections)
    if not secs:
        return str(a.session.course.cohort)
    prefixes = {s.rsplit("-", 1)[0] for s in secs}
    if len(prefixes) == 1:
        letters = sorted(s.rsplit("-", 1)[1] for s in secs)
        if letters == ["A", "B"]:
            return "AB"
        return ",".join(letters)
    return ",".join(secs)


def _spaced_day(day):
    return "    ".join(day.upper())


def _excel_col_for_slot(slot_in_day_):
    return 2 + slot_in_day_ + (1 if slot_in_day_ >= 6 else 0)


def _daily_sheets(book, assignments, rooms, cohorts, existing):
    n_slots = SLOTS_PER_DAY
    last_col = 2 + n_slots + 1  # A + 12 slots + break
    last_letter = get_column_letter(last_col)

    for day_idx, day in enumerate(DAYS):
        ws = _sheet_names(book, day.upper(), existing)

        ws.merge_cells(f"A2:{last_letter}2")
        ws["A2"] = "UNIVERSITY OF MINES AND TECHNOLOGY, TARKWA"
        ws.merge_cells(f"A3:{last_letter}3")
        ws["A3"] = "SCHOOL OF RAILWAYS AND INFRASTRUCTURE DEVELOPMENT, ESSIKADO CAMPUS"
        ws.merge_cells(f"A4:{last_letter}4")
        ws["A4"] = "SEMESTER TWO 2025/2026 TIME TABLE"

        ws.merge_cells("A6:A8")
        ws["A6"] = "CLASSROOM/\nCAPACITY"

        ws.merge_cells(f"B6:{last_letter}6")
        ws["B6"] = _spaced_day(day)

        for sl in range(n_slots):
            col = _excel_col_for_slot(sl)
            ws.cell(row=7, column=col, value=sl + 1)
            ws.cell(row=8, column=col, value=SLOT_TIMES[sl])

        ws.merge_cells("H7:H8")
        ws["H7"] = "B R E A K"

        for r in (2, 3, 4):
            cell = ws.cell(row=r, column=1)
            cell.font = Font(name="Arial", size=12, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for r in (2, 3, 4):
            ws.row_dimensions[r].height = 20
        ws.row_dimensions[6].height = 30
        ws.row_dimensions[7].height = 22
        ws.row_dimensions[8].height = 22

        cell = ws["A6"]
        cell.font = Font(name="Arial", size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c in range(1, last_col + 1):
            ws.cell(row=6, column=c).border = BORDER
            ws.cell(row=7, column=c).border = BORDER
            ws.cell(row=8, column=c).border = BORDER
            if c == 1:
                continue
            cell6 = ws.cell(row=6, column=c)
            cell6.font = Font(name="Arial", size=11)
            cell6.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=7, column=c).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=8, column=c).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=6, column=8).font = Font(name="Arial", size=9)
        ws.cell(row=7, column=8).font = Font(name="Arial", size=8)
        ws.cell(row=8, column=8).font = Font(name="Arial", size=8)
        for c in range(2, last_col + 1):
            ws.column_dimensions[get_column_letter(c)].width = 17
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["H"].width = 9

        by_room = {r.name: {} for r in rooms}
        for a in assignments:
            if day_index_of(a.slot) != day_idx:
                continue
            if not _is_no_room(a.room):
                by_room[a.room][slot_in_day(a.slot)] = a

        row = 9
        for room in rooms:
            _daily_room_row(ws, row, room.name, room.capacity, by_room[room.name])
            row += 1

        row = _daily_online_row(ws, row, day_idx, assignments)
        row = _daily_field_row(ws, row, day_idx, assignments)

        ws.freeze_panes = "B9"


def _daily_no_room_row(ws, row, day_idx, assignments, room, fill):
    from collections import defaultdict
    onl = [a for a in assignments if a.room == room and day_index_of(a.slot) == day_idx]
    cell = ws.cell(row=row, column=1, value="ONLINE (VLE)" if room == ONLINE_ROOM else "FIELD WORK")
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    cell.fill = fill

    by_slot = defaultdict(list)
    for a in onl:
        by_slot[slot_in_day(a.slot)].append(a)

    for sl in range(SLOTS_PER_DAY):
        col = _excel_col_for_slot(sl)
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = BORDER
        cell.alignment = WRAP
        classes = by_slot.get(sl)
        if classes:
            cell.value = "\n".join(
                f"{a.session.course.code} {_section_label(a)}\n{a.session.course.lecturer}"
                for a in sorted(classes, key=lambda a: a.session.course.code)
            )
    ws.row_dimensions[row].height = 60
    return row + 1


def _daily_online_row(ws, row, day_idx, assignments):
    return _daily_no_room_row(ws, row, day_idx, assignments, ONLINE_ROOM, ONLINE_FILL)


def _daily_field_row(ws, row, day_idx, assignments):
    return _daily_no_room_row(ws, row, day_idx, assignments, FIELD_WORK_ROOM, FIELD_FILL)


def _is_merged(ws, row, col):
    return any(
        r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col
        for r in ws.merged_cells.ranges
    )


def _daily_room_row(ws, row, room, capacity, by_slot):
    ws.cell(row=row, column=1, value=f"{room} ({capacity})")
    ws.cell(row=row, column=1).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=row, column=1).border = BORDER

    BREAK_COL = 8  # column H in the daily sheet
    used = set()
    for sl in range(SLOTS_PER_DAY):
        a = by_slot.get(sl)
        if a is None:
            continue
        span = max(1, a.session.duration)
        first = _excel_col_for_slot(sl)
        last = _excel_col_for_slot(min(sl + span - 1, SLOTS_PER_DAY - 1))
        if first < BREAK_COL < last:
            last = BREAK_COL - 1
        if any(c in used for c in range(first, last + 1)):
            continue
        used.update(range(first, last + 1))
        ws.merge_cells(start_row=row, start_column=first, end_row=row, end_column=last)
        c = a.session.course
        text = f"{c.code} {_section_label(a)}\n{c.lecturer}"
        cell = ws.cell(row=row, column=first, value=text)
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    break_cell = ws.cell(row=row, column=8)
    if not _is_merged(ws, row, 8):
        break_cell.fill = PatternFill("solid", fgColor="D9DEE3")
    for col in range(1, 3 + SLOTS_PER_DAY + 1):
        cell = ws.cell(row=row, column=col)
        if not _is_merged(ws, row, col):
            cell.border = BORDER
    ws.row_dimensions[row].height = 36


def _cell_text(a, include_cohort):
    c = a.session.course
    first = f"{c.code} {c.name}"
    second = f"{_room_text(a)} · {c.lecturer}"
    if include_cohort:
        second = f"{second} · {_section_label(a)}"
    return f"{first}\n{second}"


def _grid_sheet(book, base, rows_by_key, key_label, existing):
    for key in sorted(rows_by_key):
        ws = _sheet_names(book, f"{base} {key_label(key)}", existing)
        ws.cell(row=1, column=1, value="Time")
        for col, day in enumerate(DAYS, start=2):
            ws.cell(row=1, column=col, value=day)
        _style_header(ws, 1 + len(DAYS))
        for sl in range(SLOTS_PER_DAY):
            row = sl + 2
            ws.cell(row=row, column=1, value=SLOT_TIMES[sl])
            for day in range(len(DAYS)):
                text = _grid_text(rows_by_key[key], day, sl)
                if text:
                    cell = ws.cell(row=row, column=day + 2, value=text)
                    cell.alignment = WRAP
                    if _all_online(rows_by_key[key], day, sl):
                        cell.fill = ONLINE_FILL
                    elif _all_field(rows_by_key[key], day, sl):
                        cell.fill = FIELD_FILL
            ws.row_dimensions[row].height = 60
        ws.column_dimensions["A"].width = 16
        for col in range(2, 2 + len(DAYS)):
            ws.column_dimensions[get_column_letter(col)].width = 34


def _grid_text(assignments, day, slot_in_day_):
    lines = []
    for a in assignments:
        if day_index_of(a.slot) == day and slot_in_day(a.slot) == slot_in_day_:
            lines.append(_cell_text(a, False))
    return "\n".join(lines)


def _all_online(assignments, day, slot_in_day_):
    cells = [a for a in assignments if day_index_of(a.slot) == day and slot_in_day(a.slot) == slot_in_day_]
    return bool(cells) and all(a.room == ONLINE_ROOM for a in cells)


def _all_field(assignments, day, slot_in_day_):
    cells = [a for a in assignments if day_index_of(a.slot) == day and slot_in_day(a.slot) == slot_in_day_]
    return bool(cells) and all(a.room == FIELD_WORK_ROOM for a in cells)


def export_all(problem, result, output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)
    assignments = result.assignments
    cohorts = problem["cohorts"]

    book = Workbook()
    existing = set()
    book.remove(book.active)

    rooms = [r for r in problem["rooms"]]
    _daily_sheets(book, assignments, rooms, cohorts, existing)

    by_section = {sec: [] for sec in problem["sections"]}
    for a in assignments:
        for sec in a.session.sections:
            by_section[sec].append(a)

    def section_label(sec):
        return cohorts[sec].label

    _grid_sheet(book, "Cohort", by_section, section_label, existing)

    by_lecturer = {}
    for a in assignments:
        by_lecturer.setdefault(a.session.course.lecturer, []).append(a)

    _grid_sheet(book, "Lecturer", by_lecturer, lambda lec: lec, existing)

    out_path = output_dir / "timetable.xlsx"
    book.save(out_path)

    rows = []
    for a in sorted(assignments, key=lambda x: (x.slot, x.room)):
        c = a.session.course
        rows.append({
            "day": DAYS[day_index_of(a.slot)],
            "time": SLOT_TIMES[slot_in_day(a.slot)],
            "room": "ONLINE" if a.room == ONLINE_ROOM else ("FIELD WORK" if a.room == FIELD_WORK_ROOM else a.room),
            "code": c.code,
            "name": c.name,
            "programme": c.programme,
            "level": c.level,
            "cohort": _section_label(a),
            "lecturer": c.lecturer,
            "type": "Online" if a.room == ONLINE_ROOM else ("Field Work" if a.room == FIELD_WORK_ROOM else "Class"),
            "duration": a.session.duration,
        })
    (output_dir / "timetable_rows.json").write_text(json.dumps(rows, indent=2), encoding="utf-8")

    return out_path

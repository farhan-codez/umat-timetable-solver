import re
from collections import defaultdict

import openpyxl

TEACH_COLS = list(range(2, 8)) + list(range(9, 15))  # -> slots 0..5, 6..11
ONLINE_PLACEHOLDER_ROOMS = {"ONLINE", "SR 2", "SR 6", "HARDWARE LAB", "FIELD WORK", "FIELD WORK/ LAB WORK", "FIELD WORK / LAB WORK"}


def room_key(text):
    t = text.strip().upper()
    if t in ONLINE_PLACEHOLDER_ROOMS:
        return "ONLINE"
    return re.sub(r"\s*\(.*\)", "", t).strip()


def merged_map(ws):
    mm = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    mm[(r, c)] = True
    return mm


def extract_codes(text):
    tokens = text.split()
    codes = []
    pending = None
    last_end = 0
    pos = 0
    for t in tokens:
        tlen = len(t) + 1
        if re.match(r"^[A-Z/]{1,8}$", t):
            pending = t
            pos += tlen
            continue
        if pending and re.match(r"^\d{2,3}$", t):
            for L in pending.split("/"):
                codes.append((L, int(t)))
            pending = None
            last_end = pos + tlen - 1
            pos += tlen
            continue
        mg = re.match(r"^(\d{2,3})/([A-Z]{1,4})$", t)
        if mg:
            if pending:
                for L in pending.split("/"):
                    codes.append((L, int(mg.group(1))))
            pending = mg.group(2)
            pos += tlen
            continue
        break
    return codes, text[last_end:].strip()


def gen_hints(problem, ref_path):
    wb = openpyxl.load_workbook(ref_path, data_only=True)

    cells = []  # (day_idx, slot, span, room_norm, online, is_lab, code, label, paren)
    for day_idx, day in enumerate(wb.sheetnames):
        ws = wb[day]
        mm = merged_map(ws)
        cur = None
        for r in range(9, ws.max_row + 1):
            raw = ws.cell(row=r, column=1).value
            if raw and str(raw).strip():
                cur = str(raw).strip()
            if not cur:
                continue
            rk = room_key(cur)
            online_room = rk == "ONLINE"
            is_lab = not online_room and cur.strip().upper() == "COMPUTER LAB"
            for c in TEACH_COLS:
                if (r, c) in mm:
                    continue
                v = ws.cell(row=r, column=c).value
                if v is None or not str(v).strip():
                    continue
                span = 1
                for rng in ws.merged_cells.ranges:
                    if rng.min_row == r and rng.min_col == c and rng.max_row == r:
                        span = rng.max_col - rng.min_col + 1
                        break
                text = re.sub(r"\s+", " ", str(v).strip())
                codes, rest = extract_codes(text)
                if not codes:
                    continue
                online = online_room or "VLE" in text.upper()
                lab = is_lab and not online
                slot = TEACH_COLS.index(c)
                if len(codes) == 1:
                    letter, num = codes[0]
                    if not letter:
                        continue
                    code = f"{letter} {int(num):03d}"
                    label = ""
                    am = re.match(r"([AB])\b", rest)
                    if am and not re.match(r"[AB]\s*\.", rest):
                        label = "AB" if "&" in rest else am.group(1)
                    paren = None
                    pm = re.match(r"\(\s*([A-Z]{1,4})\s*([IVX]{1,5})\s*\)", rest)
                    if pm:
                        roman = pm.group(2)
                        lvl = {"I": 100, "II": 200, "III": 300, "IV": 400, "V": 500}.get(roman)
                        if lvl:
                            paren = (pm.group(1), lvl)
                    cells.append((day_idx, slot, span, rk, online, lab, code, label, paren))
    wb.close()

    room_names = {r.name for r in problem["rooms"]}
    sessions = problem["sessions"]

    has_ab = {s.course.code for s in sessions if s.course.cohort == "AB"}

    hints = []
    used_cells = set()
    for s in sessions:
        c = s.course
        dur, online, lab = s.duration, s.online, c.practical_hours > 0
        labels = [c.cohort]
        if c.cohort == "AB":
            labels.append("")
        elif c.cohort == "A" and c.code not in has_ab:
            labels.append("")
        paren = (c.programme, c.level)
        best = None
        for i, cell in enumerate(cells):
            if i in used_cells:
                continue
            if cell[2] != dur or cell[4] != online or cell[5] != lab or cell[6] != c.code:
                continue
            if cell[8]:
                if cell[8] != paren:
                    continue
            else:
                if cell[7] not in labels:
                    continue
            best = i
            break
        if best is not None:
            used_cells.add(best)
            m = cells[best]
            slot = m[0] * 12 + m[1]
            room = m[3]
            if room not in room_names:
                room = "ONLINE"
            hints.append((s, slot, room))

    sec_seen = defaultdict(list)
    conflicted_ids = set()
    for s, slot, room in hints:
        if s.online:
            continue
        for u in range(slot, slot + s.duration):
            for sec in s.sections:
                if sec_seen[(sec, u)]:
                    conflicted_ids.add(s.id)
                    for other in sec_seen[(sec, u)]:
                        conflicted_ids.add(other.id)
                sec_seen[(sec, u)].append(s)
    if conflicted_ids:
        n_before = len(hints)
        hints = [h for h in hints if h[0].id not in conflicted_ids]
        print(f"dropped {n_before - len(hints)} hints for {len(conflicted_ids)} section-conflicted sessions", flush=True)
    return hints
